#!/usr/bin/env python3
import pandas as pd
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook, Workbook
from  openpyxl.styles import Alignment, PatternFill, Font
import numpy as np
import math
import cmath
from scipy.optimize import root
from os.path import  join, abspath
#import  matplotlib as plt

# Стандартное импортирование plotly
#import plotly.plotly as py
#import plotly.graph_objs as go
#from plotly.offline import iplot

# Использование cufflinks в офлайн-режиме
import cufflinks
cufflinks.go_offline()

cracks_deep = [100, 80, 60, 40, 20]
cracks_files = []

test_deep = [100, 80, 50, 20, 15]
test_files = ['test_deep100.xlsx', 'test_deep80.xlsx', 'test_deep50.xlsx', 'test_deep20.xlsx', 'test_deep15.xlsx']

rect_files = ['rect_5_0.2_20proc.xlsx', 'rect_5_0.2_40proc.xlsx', 'rect_5_0.2_60proc.xlsx',
              'rect_5_0.2_80proc.xlsx', 'rect_5_0.2_100proc.xlsx',
              'rect_5_0.4_40_20deep.xlsx', 'rect_5_0.4_100_80_60deep.xlsx',
              'rect_5_0.6_100_80_60_40_20deep.xlsx',
              'rect_8_0.2_40_20deep.xlsx','rect_8_0.2_100_80_60deep.xlsx',
              'rect_8_0.4_40_20deep.xlsx', 'rect_8_0.4_100_80_60deep.xlsx',
              'rect_8_0.6_100_80_60_40_20deep.xlsx',
              'rect_12_0.2_40_20deep.xlsx', 'rect_12_0.2_100_80_60deep.xlsx',
              'rect_12_0.4_100_80_60_40_20deep.xlsx',
              'rect_12_0.6_100_80_60_40_20deep.xlsx']
dir = './rect'
normdir = './norm'
rectnormdir = './norm/rect_norm'

class Calibrate100(object):
	def __init__(self, z=0+0j, freq=0, norm=0+0j):
		self.freq = freq
		self.norm = norm  # на это буду домножать каждый сигнал этой частоты
		self.phase = math.degrees(cmath.phase(z))
		self.amp = abs(z)
#нормированная разность фаз для каждой частоты
		self.dphase = [] #80/50/20/15
		self.damp = []  #80/50/20/15

		self.clbr_phase = {}  # {'1':phase100, '0.8:phase80, '0.6':phase60 ...}
		self.clbr_amp = {}    # {'1':amp100, '0.8:amp80, '0.6':amp60 ...}

		self.cracks = {} #список phase для каждой Cracks
		self.dphase_cracks = {} #список dphase для каждой Cracks
		#{
		#   (5, 0.2): {'1':phase100, '0.8:phase80, '0.6':phase60 ...},
		#   (5, 0.4):  {'1':phase100, '0.8:phase80, '0.6':phase60 ...},
		#   (5, 0.6): {'1':phase100, '0.8:phase80, '0.6':phase60 ...},
		#     ...
		#   (12, 0.8):  {'1':phase100, '0.8:phase80, '0.6':phase60 ...},
		#   (12, 1): {'1':phase100, '0.8:phase80, '0.6':phase60 ...}
		# }

class ThisCracks(object):
	def __init__(self, crack_type=None,  crack_len=0, crack_w=0, deep=0, freq=0):

		self.freq = freq
		self.type = crack_type
		self.deep = deep

		self.len = crack_len
		self.w = crack_w

		self.phase =  0
		self.amp = 0
		self.data = []


def fun(x):
	return [x[0]**2 + x[1]**2 - 100.0,
		x[0]*math.tan(math.radians(40))-x[1]]


def normalize(z):
	A = abs(z)
	fi = math.degrees(cmath.phase(z))
	print(f'normalize:\nA_orig = {A}, fi_orig = {fi}')

	#поиск im/re, соответсвующих амплитуде 10В и фазе 40 градусов
	sol=root(fun, [0, 0])
	z_et = -1*complex(sol.x[0], sol.x[1])
	#norm - нормировочный коэффицент
	norm = z_et/z

	print(f'z_norm = {z*norm}\nAnorm = {abs(z*norm)}, fi = {math.degrees(cmath.phase(z*norm))}')
	#вернуть нормировочный коэффициент
	return(norm)


def fill_nested_dict(ws, start_row):
	freq_dict = {}
	for row in ws.iter_rows(min_row=start_row, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
		if len(row) > 0:
			freq = row[2].value
			if freq is not None:
				if freq not in freq_dict:
					freq_dict[freq] = {}
			# print(f'empty freq_dict: {freq_dict}')
			deep = row[1].value
			if deep is not None:
				data = [cell.value for cell in row]
				if deep not in freq_dict[freq]:
					freq_dict[freq][deep] = []
				freq_dict[freq][deep].append(data)
	return (freq_dict)


def fill_dict(ws, start_row):
	mandata = {}
	for row in ws.iter_rows(min_row=start_row, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
		if len(row) > 0:
			freq = row[1].value
			if freq is not None:
				data = [cell.value for cell in row]
				#print(data)
				if freq not in mandata:
					mandata[freq] = []
				mandata[freq].append(data)
	return (mandata)


def fill_refdata(clbrt_dict):
	# lambda-функции для сортировки по ключу
	ampl = lambda data: data[4]
	def_disp = lambda data: data[0]
#   [0]             [1]         [2]         [3]     [4]
#['def_disp[mm]', 'freq[kHz]', 'Im[V]', 'Re[V]', 'Amplitude[V]']
	Calibrate_Curve = []
	for freq in clbrt_dict:
		# sorting by Amplitude[V](maxZ)
		clbrt_dict[freq].sort(key=ampl, reverse=True)

		# complex number formation for turning
		z = complex(clbrt_dict[freq][0][3], clbrt_dict[freq][0][2])
		norm = normalize(z)
		print(f'freq = {freq},norm = {norm}')

		clbrt_data = Calibrate100(complex((z * norm).real, (z * norm).imag),
		                          int(clbrt_dict[freq][0][1]), norm)
		clbrt_data.clbr_phase = {}
		clbrt_data.clbr_phase[1.0] = math.degrees(cmath.phase(z*norm))
		clbrt_data.clbr_amp[1.0] = abs(z*norm)
		print(f'clbrt_data.clbr_phase= {clbrt_data.clbr_phase}')
		Calibrate_Curve.append(clbrt_data)
		for row in clbrt_dict[freq]:
			z = complex(row[3], row[2])
			row[3] = (z * norm).real
			row[2] = (z * norm).imag
			row[4] = abs(z * norm)
		# sorting by def_disp[mm]
		clbrt_dict[freq].sort(key=def_disp, reverse=False)
		print('*******\n')
	return Calibrate_Curve

def save_alone_deep(dict,header, name):
	for freq in dict:
		exname = (name.split('.'))[0] + '_' + str(int(freq)) + 'kHz_norm.xlsx'
		file_path = join(normdir, exname)
		if not os.path.exists(file_path):
			wb = Workbook()
			ws = wb.active
			ws.append(header)
			for row in dict[freq]:
				ws.append(row)
			# выравнивание колонок:
			for i in range(1, 5):
				zagl = ws.cell(row=1, column=i)
				zagl.alignment = Alignment(horizontal='center')
			wb.save(file_path)

def save_cracks(dict, file_path, header, num_col):
	wb = Workbook()
	ws = wb.active
	ws.append(header)
	for row in dict:
		ws.append(row)
	# выравнивание колонок:
	for i in range(1, num_col):
		zagl = ws.cell(row=1, column=i)
		zagl.alignment = Alignment(horizontal='center')
	wb.save(file_path)

def main():

	file_path = join(dir, test_files[0])
	wb = load_workbook(filename=file_path, data_only=True, read_only=True)
	wsn = list(wb.sheetnames)
	ws = wb.active
	header = [cell.value for cell in next(
			ws.iter_rows(min_row=1, min_col=1, max_col=ws.max_column))]
	print(header)
	wb.close

# обработка сигнала от сквозного отверстия:
# получение нормировочного коэффициента для каждой частоты
# создание листа Calibrate_Curve опорных данных от сквозного отверстия

# составление dict с частотой в качестве ключа и вложенным списком строк данных в качестве значения
# {
# '25' : [[row[0]],[row[1]],..,[row[len(mandata[freq])-1]]],
# '100': [[row[0]],[row[1]],..,[row[len(mandata[freq])-1]]],
# '200': [[row[0]],[row[1]],..,[row[len(mandata[freq])-1]]]
# }
	clbrt_dict= fill_dict(ws, start_row=2)
	Calibrate_Curve = fill_refdata(clbrt_dict)

	# создание файлов нормированных калибровочных сквозных сигналов
	for freq in clbrt_dict:
		exname = 'test_deep100_' + str(int(freq)) + 'kHz_norm.xlsx'
		file_path = join(normdir, exname)
		if not os.path.exists(file_path):
			wb = Workbook()
			ws = wb.active
			ws.append(header)
			for row in clbrt_dict[freq]:
				ws.append(row)
			#выравнивание колонок:
			for i in range(1, 5):
				zagl = ws.cell(row = 1, column=i)
				zagl.alignment = Alignment(horizontal='center')
			wb.save(file_path)
#*************************************************************************

# сортировка и вычисление dphase & damp для остальных калибровочных отверстий
# запись нормированных данных в файлы в папку ./norm
	for i in range(1, len(test_files)):
		name = test_files[i]
		test_split = name.split('_')
		file_path = join(dir, test_files[i])
		print(file_path)
		deep = (int(test_split[1][4]) * 10 + int(test_split[1][5]))/100
		wb = load_workbook(filename=file_path, data_only=True, read_only=True)
		ws = wb.active
		# сохранение заголовка для последующего сохранения в новый файл
		header = [cell.value for cell in next(
			ws.iter_rows(min_row=1, min_col=1, max_col=ws.max_column))]
		wb.close

		dict=fill_dict(ws=ws, start_row=2)

		ampl = lambda data: data[4]
		def_disp = lambda data: data[0]

		for freq in dict:
			print(f'freq = {freq}')
			for n in range(0, len(Calibrate_Curve)):
				if Calibrate_Curve[n].freq == freq:
					norm = Calibrate_Curve[n].norm

			#print(f'n = {n}, freq = {freq}, norm = {norm}')
			for row in dict[freq]:
				z = complex(row[3], row[2])
				row[3] = (z * norm).real
				row[2] = (z * norm).imag
				row[4] = abs(z * norm)

		# сохранение опорных данных  фазы и макс амплидуды
			# sorting by Amplitude[V](maxZ)
			dict[freq].sort(key=ampl, reverse=True)
			z = complex(dict[freq][0][3], dict[freq][0][2])

			for n in range(0, len(Calibrate_Curve)):
				if Calibrate_Curve[n].freq == freq:
					Calibrate_Curve[n].clbr_phase[deep] = math.degrees(cmath.phase(z))
					Calibrate_Curve[n].clbr_amp[deep] = abs(z)

					Calibrate_Curve[n].dphase.append(math.degrees(cmath.phase(z)) - Calibrate_Curve[n].phase)
					Calibrate_Curve[n].damp.append(Calibrate_Curve[n].amp / abs(z))
					print(f'Phase = {math.degrees(cmath.phase(z))}, dphase = {Calibrate_Curve[n].dphase[i-1]}')
					print(f'Amp = {abs(z)}, damp = {Calibrate_Curve[n].damp[i - 1]}')
					print(f'clbrt_phase = {Calibrate_Curve[n].clbr_phase}')
			# возврат сигнала в исходное положение
			# sorting by def_disp[mm]
			dict[freq].sort(key=def_disp, reverse=False)
			print(dict[freq])
			save_alone_deep(dict, header, name)
			print('********************************\n')

	#*********************************************************************

	# обработка сигналов от прямоугольных трещин
	# запись нормированных данных в файлы в папку ./norm
	RectCracks = []# хранятся данные о каждой трещине для записи в файл и парсинга разницы фаз

	for i in range(0, len(rect_files)):
		name = rect_files[i]
		file_path = join(dir, name)
		#print(name)

		crack_split = name.split('_')
		crack_type = crack_split[0]
		crack_len = int(crack_split[1])
		crack_w = float(crack_split[2])

		wb = load_workbook(filename=file_path, data_only=True, read_only=True)
		ws = wb.active
		header = [cell.value for cell in next(
			ws.iter_rows(min_row=1, min_col=1, max_col=ws.max_column))]
		wb.close

		# составление dict с частотой в качестве ключа и вложенным списком строк данных в качестве значения
		# {
		# '25' : {'1': [[row[0]],[row[1]],..,[row[len(mandata[freq])-1]]]},
		#        {'0.8': [[row[0]],[row[1]],..,[row[len(mandata[freq])-1]]]},
		#        {'0.6': [[row[0]],[row[1]],..,[row[len(mandata[freq])-1]]]},
		#        {'0.4': [[row[0]],[row[1]],..,[row[len(mandata[freq])-1]]]},
		#....
		# '100': [[row[0]],[row[1]],..,[row[len(mandata[freq])-1]]],
		# '200': [[row[0]],[row[1]],..,[row[len(mandata[freq])-1]]]
		# }
		if (header[1] == 'freq[kHz]') or (header[1] == 'freq (kHz)'):
			freq_dict = fill_dict(ws, start_row=2)
			for freq in freq_dict:
				deep = 0
				for n in crack_split[3]:
					if n.isdigit():
						deep = deep*10 + int(n)
				crack_data = ThisCracks(crack_type, crack_len, crack_w,
				                        float(deep)/100, freq)
				for n in range(0, len(Calibrate_Curve)):
					if Calibrate_Curve[n].freq == freq:
						norm = Calibrate_Curve[n].norm
				for row in freq_dict[freq]:
					z = complex(float(row[3]), float(row[2]))
					row[3] = (z * norm).real
					row[2] = (z * norm).imag
					row[4] = abs(z * norm)

				ampl = lambda data: float(data[4])
				freq_dict[freq].sort(key=ampl, reverse=False)
				z = complex(freq_dict[freq][0][3], freq_dict[freq][0][2])
				crack_data.phase = math.degrees(cmath.phase(z))
				crack_data.amp = abs(z)

				# sorting by def_disp[mm]
				def_disp = lambda data: float(data[0])
				freq_dict[freq].sort(key=def_disp, reverse=False)
				crack_data.data = freq_dict[freq]
				RectCracks.append(crack_data)
				exname = crack_type + '_' + str(int(crack_len)) + '_' + str(float(crack_w)) + '_' + str(int(deep)) + 'deep' + '_' + str(int(freq)) + 'kHz_norm.xlsx'
				new_file_path = join(rectnormdir, exname)
				if not os.path.exists(new_file_path):
					save_cracks(dict=freq_dict[freq], file_path=new_file_path, header=header, num_col=5)
		else:
			freq_dict = fill_nested_dict(ws, 2)
			for freq in freq_dict:
				for deep in freq_dict[freq]:
					crack_data = ThisCracks(crack_type, crack_len, crack_w, deep, freq)
					for n in range(0,len(Calibrate_Curve)):
						if Calibrate_Curve[n].freq == freq:
							norm = Calibrate_Curve[n].norm
					for row in freq_dict[freq][deep]:
						z = complex(float(row[4]), float(row[3]))
						row[4] = (z * norm).real
						row[3] = (z * norm).imag
						row[5] = abs(z * norm)

					ampl = lambda data: float(data[5])
					freq_dict[freq][deep].sort(key=ampl, reverse=False)
					z = complex(freq_dict[freq][deep][0][4], freq_dict[freq][deep][0][3])
					crack_data.phase = math.degrees(cmath.phase(z))
					crack_data.amp = abs(z)

					# возврат сигнала в исходное положение
					# sorting by def_disp[mm]
					def_disp = lambda data: float(data[0])
					freq_dict[freq][deep].sort(key=def_disp, reverse=False)
					crack_data.data = freq_dict[freq][deep]
					RectCracks.append(crack_data)

					#записать RectCracks в файлы
					exname = crack_type + '_'+str(int(crack_len))+'_' + str(float(crack_w))+'_'+str(int(float(deep)*100))+'deep'+'_'+ str(int(freq)) + 'kHz_norm.xlsx'
					new_file_path = join(rectnormdir, exname)
					if not os.path.exists(new_file_path):
						save_cracks(dict=freq_dict[freq][deep], file_path=new_file_path, header=header, num_col=6)

	#объединение для  Calibrsate_Curve для трещины c конкретными параметрами len, w на каждой частоте
	for cc in Calibrate_Curve:
		for rc in RectCracks:
			if cc.freq == rc.freq:
				tup = tuple((rc.len, rc.w))
				if tup not in cc.cracks:
					cc.cracks[tup] = {}
				if rc.deep not in cc.cracks[tup]:
					cc.cracks[tup][rc.deep] = rc.phase
		print(f'freq ={cc.freq}, cracks = {cc.cracks}')

	#print('******\n')
	for cc in Calibrate_Curve:
		for crack in cc.cracks:
			if crack not in cc.dphase_cracks:
				cc.dphase_cracks[crack] = {}
			for deep in cc.cracks[crack]:
				if deep not in cc.dphase_cracks[crack]:
					cc.dphase_cracks[crack][deep] = cc.cracks[crack][deep] - cc.phase
	print('\n')
	for cc in Calibrate_Curve:
		print(f'freq = {cc.freq}, clbr_phase = {cc.clbr_phase}')

	print('Marina very clever!\n')

#********************VIS**************************************
	#import plotly.plotly as py
	from plotly import graph_objs as go
	#import plotly.express as px


	for cc in Calibrate_Curve:
		title = f'freq = {cc.freq}'

		fig = go.Figure()
		deep = []
		deg = []
		sorted_by_value = sorted(cc.clbr_phase.items(), key=lambda kv: kv[1], reverse=False)
		for i in sorted_by_value:
			deep.append(int(float(i[0]) * 100))
			deg.append(i[1])
		fig.add_trace(go.Scatter(x=deg, y=deep, name='Calibrate' + title, line_shape='spline'))

		for crack in cc.cracks:
			#print(crack)
			crack_deep = []
			crack_deg = []
			sort = sorted(cc.cracks[crack].items(), key=lambda kv: kv[1], reverse=False)
			#print(sort)
			for i in sort:
				if (crack[0] == 5):
					crack_deep.append(int(float(i[0]) * 100))
					crack_deg.append(i[1])
			fig.add_trace(go.Scatter(x=crack_deg, y=crack_deep, text=[str(i)], mode='markers', name=str(crack)))

		fig.update_layout(title_text=title, xaxis_type="linear", yaxis_type="linear",
		                  xaxis_title="Phase [degrees]",
		                  yaxis_title="Deep [%]")
		fig.show()



'''
	var dotDiv = document.getElementById("dot-chart");
 
var traceA = {
  type: "scatter",
  mode: "markers",
  x: [2011, 2012, 2013, 2014, 2015, 2016],
  y: [789, 795, 760, 775, 780, 783],
  name: 'Highest Marks',
  marker: {
    color: 'rgba(156, 165, 196, 0.5)',
    line: {
      color: 'rgba(156, 165, 196, 1)',
      width: 1,
    },
    symbol: 'circle',
    size: 20
  },
  hoverlabel: {
    bgcolor: 'black',
  }
};
 
var traceB = {
  type: "scatter",
  mode: "markers",
  x: [2011, 2012, 2013, 2014, 2015, 2016],
  y: [769, 755, 747, 770, 771, 781],
  name: 'Second Highest Marks',
  marker: {
    color: 'rgba(165, 196, 50, 0.5)',
    line: {
      color: 'rgba(165, 196, 50, 1)',
      width: 1,
    },
    symbol: 'circle',
    size: 20
  },
  hoverlabel: {
    bgcolor: 'black',
  }
};
 
var data = [traceA, traceB];
 
var layout = {
  title: 'Marks Obtained by Top Two Students',
  xaxis: {
    showgrid: false,
    showline: true,
    linecolor: 'rgb(200, 0, 0)',
    ticks: 'outside',
    tickcolor: 'rgb(200, 0, 0)',
    tickwidth: 4
  },
  legend: {
    bgcolor: 'white',
    borderwidth: 1,
    bordercolor: 'black',
    orientation: 'h',
    xanchor: 'center',
    x: 0.5,
    font: {
      size: 12,
    }
  },
  paper_bgcolor: 'rgb(255, 230, 255)',
  plot_bgcolor: 'rgb(255, 230, 255)'
};
 
Plotly.plot(dotDiv, data, layout);'''

if __name__ == "__main__":
	main()

'''movies = pd.read_excel(file)
print(movies.head())
#sorted_tab = movies.sort_values(by=['freq[kHz]', 'deep', 'def_disp[mm]'], ascending=True)
sorted_tab = movies.sort_values(by=['freq[kHz]', 'def_disp[mm]'], ascending=True)
print(sorted_tab.head())
#print(sorted_tab[lambda x: x['freq[kHz]'] == 25])

#print(sorted_tab.index)
#print(sorted_tab.columns)
print(sorted_tab.values)

#for i in range (sorted_tab.index):
	#if (sorted_tab[lambda x: x['freq[kHz]'] == 25]):

#print(sorted_tab['freq[kHz]'].head())

#interpolated =sorted_tab.head(33).interpolate(method='polynomial', order=5)
#interpolated.plot('Re[V]', 'Im[V]')
#sorted_tab.head(62).plot('Re[V]','Im[V]', kind='density')
#plt.show()
print(movies.shape)'''




